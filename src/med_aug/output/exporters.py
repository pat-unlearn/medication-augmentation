"""Data exporters for various formats."""

from typing import Dict, Any, List, Optional
from pathlib import Path
import json
import csv
from datetime import datetime
from abc import ABC, abstractmethod

from ..core.logging import get_logger

logger = get_logger(__name__)


class BaseExporter(ABC):
    """Base class for data exporters."""

    @abstractmethod
    def export(self, data: Any, output_path: Path, **kwargs) -> Path:
        """
        Export data to file.

        Args:
            data: Data to export
            output_path: Output file path
            **kwargs: Additional options

        Returns:
            Path to exported file
        """
        pass


class JSONExporter(BaseExporter):
    """Export data to JSON format."""

    def export(self, data: Any, output_path: Path, indent: int = 2, **kwargs) -> Path:
        """Export to JSON."""
        logger.info("json_export_started", path=str(output_path))

        # Ensure .json extension
        if output_path.suffix != ".json":
            output_path = output_path.with_suffix(".json")

        # Convert data if needed
        export_data = self._prepare_data(data)

        # Write JSON
        with open(output_path, "w") as f:
            json.dump(export_data, f, indent=indent, default=str)

        logger.info(
            "json_export_completed",
            path=str(output_path),
            size_bytes=output_path.stat().st_size,
        )

        return output_path

    def _prepare_data(self, data: Any) -> Any:
        """Prepare data for JSON serialization."""
        if hasattr(data, "to_dict"):
            return data.to_dict()
        elif hasattr(data, "__dict__"):
            return data.__dict__
        return data


class CSVExporter(BaseExporter):
    """Export data to CSV format."""

    def export(self, data: List[Dict[str, Any]], output_path: Path, **kwargs) -> Path:
        """Export to CSV."""
        logger.info("csv_export_started", path=str(output_path))

        # Ensure .csv extension
        if output_path.suffix != ".csv":
            output_path = output_path.with_suffix(".csv")

        if not data:
            logger.warning("csv_export_empty_data")
            with open(output_path, "w") as f:
                f.write("")
            return output_path

        # Get fieldnames
        fieldnames = list(data[0].keys())

        # Write CSV
        with open(output_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)

        logger.info(
            "csv_export_completed",
            path=str(output_path),
            rows=len(data),
            columns=len(fieldnames),
        )

        return output_path


class ExcelExporter(BaseExporter):
    """Export data to Excel format with advanced formatting."""

    def export(self, data: Dict[str, Any], output_path: Path, **kwargs) -> Path:
        """
        Export to Excel with multiple sheets.

        Args:
            data: Dictionary mapping sheet names to data
            output_path: Output file path

        Returns:
            Path to exported file
        """
        import pandas as pd

        logger.info("excel_export_started", path=str(output_path))

        # Ensure .xlsx extension
        if output_path.suffix not in [".xlsx", ".xls"]:
            output_path = output_path.with_suffix(".xlsx")

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Process each sheet
            for sheet_name, sheet_data in data.items():
                if isinstance(sheet_data, list) and sheet_data:
                    # List of dictionaries -> DataFrame
                    df = pd.DataFrame(sheet_data)
                elif isinstance(sheet_data, dict):
                    # Dictionary -> DataFrame (transpose)
                    df = pd.DataFrame([sheet_data])
                elif hasattr(sheet_data, "to_frame"):
                    # Series -> DataFrame
                    df = sheet_data.to_frame()
                elif isinstance(sheet_data, pd.DataFrame):
                    df = sheet_data
                else:
                    logger.warning(
                        "excel_export_unsupported_data_type",
                        sheet=sheet_name,
                        data_type=type(sheet_data).__name__,
                    )
                    continue

                # Write to Excel
                df.to_excel(
                    writer, sheet_name=sheet_name[:31], index=False
                )  # Excel sheet name limit

                # Get worksheet for formatting
                worksheet = writer.sheets[sheet_name[:31]]

                # Apply formatting
                self._format_worksheet(worksheet, df)

            # Add metadata sheet
            metadata = {
                "Generated": datetime.now().isoformat(),
                "Generator": "Medication Augmentation System",
                "Version": "1.0.0",
                "Sheets": list(data.keys()),
            }
            pd.DataFrame([metadata]).to_excel(
                writer, sheet_name="_Metadata", index=False
            )

        logger.info("excel_export_completed", path=str(output_path), sheets=len(data))

        return output_path

    def _format_worksheet(self, worksheet, dataframe):
        """Apply formatting to worksheet."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Format data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = min(length + 2, 50)

        # Freeze top row
        worksheet.freeze_panes = "A2"


class HTMLExporter(BaseExporter):
    """Export data to HTML format."""

    def export(
        self,
        data: Dict[str, Any],
        output_path: Path,
        title: str = "Data Export",
        **kwargs,
    ) -> Path:
        """Export to HTML."""
        logger.info("html_export_started", path=str(output_path))

        # Ensure .html extension
        if output_path.suffix != ".html":
            output_path = output_path.with_suffix(".html")

        # Generate HTML
        html_content = self._generate_html(data, title)

        # Write file
        with open(output_path, "w") as f:
            f.write(html_content)

        logger.info("html_export_completed", path=str(output_path))

        return output_path

    def _generate_html(self, data: Dict[str, Any], title: str) -> str:
        """Generate HTML content."""
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #4CAF50;
            padding-bottom: 10px;
        }}
        h2 {{
            color: #555;
            margin-top: 30px;
            border-bottom: 1px solid #ddd;
            padding-bottom: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th {{
            background: #4CAF50;
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background: #f9f9f9;
        }}
        tr:hover {{
            background: #f5f5f5;
        }}
        .metric {{
            display: inline-block;
            margin: 10px 20px 10px 0;
            padding: 10px;
            background: #f0f0f0;
            border-radius: 4px;
        }}
        .metric-label {{
            font-weight: bold;
            color: #666;
        }}
        .metric-value {{
            color: #4CAF50;
            font-size: 1.2em;
            margin-left: 10px;
        }}
        .timestamp {{
            color: #999;
            font-size: 0.9em;
            text-align: right;
            margin-top: 20px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        """

        # Add content sections
        for section_name, section_data in data.items():
            html += f"<h2>{section_name}</h2>"

            if (
                isinstance(section_data, list)
                and section_data
                and isinstance(section_data[0], dict)
            ):
                # Render as table
                html += "<table>"
                html += "<thead><tr>"
                for key in section_data[0].keys():
                    html += f"<th>{key}</th>"
                html += "</tr></thead><tbody>"

                for row in section_data[:100]:  # Limit to 100 rows
                    html += "<tr>"
                    for key in section_data[0].keys():
                        value = row.get(key, "")
                        html += f"<td>{value}</td>"
                    html += "</tr>"

                html += "</tbody></table>"

                if len(section_data) > 100:
                    html += (
                        f"<p><em>Showing first 100 of {len(section_data)} rows</em></p>"
                    )

            elif isinstance(section_data, dict):
                # Render as metrics
                for key, value in section_data.items():
                    html += f"""
                    <div class="metric">
                        <span class="metric-label">{key}:</span>
                        <span class="metric-value">{value}</span>
                    </div>
                    """
            else:
                # Render as text
                html += f"<div>{section_data}</div>"

        # Add timestamp
        html += f"""
        <div class="timestamp">Generated: {datetime.now().isoformat()}</div>
    </div>
</body>
</html>
        """

        return html


class PDFExporter(BaseExporter):
    """Export data to PDF format."""

    def export(self, data: Any, output_path: Path, **kwargs) -> Path:
        """Export to PDF."""
        logger.info("pdf_export_started", path=str(output_path))

        # Ensure .pdf extension
        if output_path.suffix != ".pdf":
            output_path = output_path.with_suffix(".pdf")

        try:
            # Try using reportlab if available
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import (
                SimpleDocTemplate,
                Table,
                TableStyle,
                Paragraph,
                Spacer,
            )
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors

            # Create PDF document
            doc = SimpleDocTemplate(str(output_path), pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Add title
            title = kwargs.get("title", "Medication Augmentation Report")
            elements.append(Paragraph(title, styles["Title"]))
            elements.append(Spacer(1, 20))

            # Add content based on data type
            if isinstance(data, dict):
                for section_name, section_data in data.items():
                    elements.append(Paragraph(section_name, styles["Heading2"]))
                    elements.append(Spacer(1, 10))

                    if isinstance(section_data, list) and section_data:
                        # Create table
                        table_data = []
                        if isinstance(section_data[0], dict):
                            # Add headers
                            table_data.append(list(section_data[0].keys()))
                            # Add rows
                            for row in section_data[:50]:  # Limit rows
                                table_data.append(
                                    [
                                        str(row.get(k, ""))
                                        for k in section_data[0].keys()
                                    ]
                                )
                        else:
                            table_data = [[str(item)] for item in section_data[:50]]

                        table = Table(table_data)
                        table.setStyle(
                            TableStyle(
                                [
                                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                                ]
                            )
                        )
                        elements.append(table)

                    elements.append(Spacer(1, 20))

            # Build PDF
            doc.build(elements)

            logger.info("pdf_export_completed", path=str(output_path))
            return output_path

        except ImportError:
            logger.warning("pdf_export_unavailable", reason="reportlab not installed")

            # Fallback to HTML
            html_exporter = HTMLExporter()
            html_path = output_path.with_suffix(".html")
            return html_exporter.export(data, html_path, **kwargs)


class MarkdownExporter(BaseExporter):
    """Export data to Markdown format."""

    def export(
        self,
        data: Dict[str, Any],
        output_path: Path,
        title: str = "Medication Augmentation Report",
        **kwargs,
    ) -> Path:
        """Export to Markdown."""
        logger.info("markdown_export_started", path=str(output_path))

        # Ensure .md extension
        if output_path.suffix != ".md":
            output_path = output_path.with_suffix(".md")

        # Generate Markdown
        md_content = self._generate_markdown(data, title)

        # Write file
        with open(output_path, "w") as f:
            f.write(md_content)

        logger.info("markdown_export_completed", path=str(output_path))

        return output_path

    def _generate_markdown(self, data: Dict[str, Any], title: str) -> str:
        """Generate Markdown content."""
        lines = [
            f"# {title}",
            "",
            f"*Generated: {datetime.now().isoformat()}*",
            "",
            "---",
            "",
        ]

        for section_name, section_data in data.items():
            lines.append(f"## {section_name}")
            lines.append("")

            if isinstance(section_data, list) and section_data:
                if isinstance(section_data[0], dict):
                    # Table format
                    headers = list(section_data[0].keys())
                    lines.append("| " + " | ".join(headers) + " |")
                    lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

                    for row in section_data[:50]:
                        values = [str(row.get(h, "")) for h in headers]
                        lines.append("| " + " | ".join(values) + " |")
                else:
                    # List format
                    for item in section_data[:50]:
                        lines.append(f"- {item}")

                if len(section_data) > 50:
                    lines.append("")
                    lines.append(f"*Showing first 50 of {len(section_data)} items*")

            elif isinstance(section_data, dict):
                # Metrics format
                for key, value in section_data.items():
                    lines.append(f"- **{key}**: {value}")

            else:
                # Text format
                lines.append(str(section_data))

            lines.append("")

        return "\n".join(lines)


class ConmedsYAMLExporter(BaseExporter):
    """Export data to conmeds.yml format for clinical pipeline integration."""

    def export(
        self,
        data: Dict[str, Any],
        output_path: Path,
        title: str = "Medication Augmentation - NSCLC",
        **kwargs,
    ) -> Path:
        """
        Export to conmeds.yml format.

        Args:
            data: Dictionary with 'classifications' key containing medication classifications
            output_path: Output file path
            title: Title for the YAML file
            **kwargs: Additional options including 'conmeds_file' for base conmeds file

        Returns:
            Path to exported conmeds.yml file
        """

        logger.info("conmeds_yaml_export_started", path=str(output_path))

        # Ensure .yml extension
        if output_path.suffix not in [".yml", ".yaml"]:
            output_path = output_path.with_suffix(".yml")

        # Load existing conmeds as base
        base_conmeds = self._load_base_conmeds(kwargs.get("conmeds_file"))

        # Extract classifications from data
        classifications = data.get("classifications", {})
        if not classifications:
            # Try alternative data structures
            if "llm_classifications" in data:
                classifications = data["llm_classifications"].get("classifications", {})
            elif "classification_results" in data:
                # Convert individual results to classifications dict
                classifications = {}
                for result in data["classification_results"]:
                    drug_class = result.get("drug_class")
                    medication = result.get("medication")
                    if drug_class and medication:
                        if drug_class not in classifications:
                            classifications[drug_class] = []
                        classifications[drug_class].append(medication)

        # Augment base conmeds with new classifications
        augmented_conmeds = self._augment_conmeds(base_conmeds, classifications)

        # Add metadata as comments
        yaml_content = self._generate_yaml_with_metadata(augmented_conmeds, title)

        # Write YAML file
        with open(output_path, "w") as f:
            f.write(yaml_content)

        logger.info(
            "conmeds_yaml_export_completed",
            path=str(output_path),
            drug_classes=len(augmented_conmeds),
            total_medications=sum(len(meds) for meds in augmented_conmeds.values()),
        )

        return output_path

    def _load_base_conmeds(self, conmeds_file: Optional[str]) -> Dict[str, List[str]]:
        """
        Load existing conmeds file as the base for augmentation.

        Args:
            conmeds_file: Path to the base conmeds file

        Returns:
            Dict in conmeds.yml format with 'taking_drugname' keys
        """
        import yaml

        if not conmeds_file:
            logger.warning(
                "no_base_conmeds_file",
                message="No base conmeds file provided, starting empty",
            )
            return {}

        conmeds_path = Path(conmeds_file)
        if not conmeds_path.exists():
            logger.warning("base_conmeds_not_found", path=str(conmeds_path))
            return {}

        try:
            with open(conmeds_path, "r") as f:
                base_conmeds = yaml.safe_load(f)

            if not isinstance(base_conmeds, dict):
                logger.warning("invalid_base_conmeds_format", path=str(conmeds_path))
                return {}

            logger.info(
                "base_conmeds_loaded",
                path=str(conmeds_path),
                drug_classes=len(base_conmeds),
                total_medications=sum(
                    len(meds) if isinstance(meds, list) else 0
                    for meds in base_conmeds.values()
                ),
            )

            return base_conmeds
        except Exception as e:
            logger.error(
                "base_conmeds_load_error", path=str(conmeds_path), error=str(e)
            )
            return {}

    def _augment_conmeds(
        self, base_conmeds: Dict[str, List[str]], classifications: Dict[str, List[str]]
    ) -> Dict[str, List[str]]:
        """
        Augment base conmeds with new medication classifications.

        Args:
            base_conmeds: Existing conmeds in simple array format
            classifications: New classifications to add

        Returns:
            Dict in conmeds.yml format with augmented medications
        """
        # Start with a copy of the base conmeds
        augmented = base_conmeds.copy()

        for drug_class, medications in classifications.items():
            # Convert drug class name to conmeds format
            conmeds_key = self._convert_drug_class_name(drug_class)

            # If this drug class already exists, add to it
            if conmeds_key in augmented:
                existing_meds = set(augmented[conmeds_key])
                new_meds = [med for med in medications if med not in existing_meds]
                if new_meds:
                    augmented[conmeds_key].extend(new_meds)
                    # Keep sorted and unique
                    augmented[conmeds_key] = sorted(list(set(augmented[conmeds_key])))
            else:
                # New drug class - add it
                unique_meds = sorted(list(set(medications)))
                augmented[conmeds_key] = unique_meds

        return augmented

    def _convert_drug_class_name(self, drug_class: str) -> str:
        """
        Convert drug class name to conmeds.yml format.

        Examples:
            "chemotherapy" -> "taking_chemotherapy"
            "egfr_inhibitors" -> "taking_egfr_inhibitors"
            "pembrolizumab" -> "taking_pembrolizumab"
        """
        # Normalize the name
        normalized = drug_class.lower().strip()

        # Remove common prefixes if present
        prefixes_to_remove = ["taking_", "drug_class_", "class_"]
        for prefix in prefixes_to_remove:
            if normalized.startswith(prefix):
                normalized = normalized[len(prefix) :]

        # Handle specific drug names vs. drug classes
        # If it's a specific drug name, use it directly
        # If it's a drug class, keep the class name

        return f"taking_{normalized}"

    def _generate_yaml_with_metadata(
        self, conmeds: Dict[str, List[str]], title: str
    ) -> str:
        """Generate YAML content matching original format exactly."""
        lines = []

        # Sort keys to match original order
        for key in sorted(conmeds.keys()):
            medications = conmeds[key]
            # Format as simple array on single line like original
            med_list = ", ".join(medications)
            lines.append(f"{key}: [{med_list}]")

        return "\n".join(lines) + "\n"
